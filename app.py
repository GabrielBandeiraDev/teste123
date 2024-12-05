from flask import Flask, render_template, request, redirect, jsonify, session
import json
import os
import apidomercadopago
from apidomercadopago import gerar_link_pagamento
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = os.urandom(24)  # Necessário para usar a sessão

# Caminho para o arquivo JSON e Excel
JSON_FILE = "pagamentos.json"
EXCEL_FILE = "clientes.xlsx"

# Função para inicializar o arquivo JSON, caso não exista
def inicializar_json():
    if not os.path.exists(JSON_FILE):
        with open(JSON_FILE, "w") as f:
            json.dump({"contador_pagamentos": 0}, f)

# Função para atualizar o contador de pagamentos
def registrar_pagamento():
    try:
        with open(JSON_FILE, "r+") as f:
            data = json.load(f)
            data["contador_pagamentos"] += 1
            f.seek(0)
            json.dump(data, f, indent=4)
            f.truncate()  # Remove o excesso de dados antigos
    except Exception as e:
        print(f"Erro ao atualizar o JSON: {e}")

# Função para salvar os dados no Excel
def salvar_no_excel(nome, email, telefone, evento):
    # Verifica se o arquivo Excel já existe
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Nome", "Email", "Telefone", "Evento"])  # Cabeçalho
    
    # Adiciona os dados do cliente
    sheet.append([nome, email, telefone, evento])
    wb.save(EXCEL_FILE)








# Rota da homepage
@app.route("/")
def homepage():
    return render_template("index.html")

# Demais rotas
@app.route("/escolha")
def escolha():
    return render_template("Escolha.html")



@app.route("/escolhaconv")
def escolhaconv():
    return render_template("escolhaconv.html")

@app.route("/comeca")
def comeca():
    return render_template("comeca.html")

@app.route("/igreja")
def igreja():
    return render_template("igreja.html")














@app.route("/inscricao", methods=["GET", "POST"])
def inscricao():
    if request.method == "POST":
        nome = request.form["nome"]
        email = request.form["email"]
        telefone = request.form["telefone"]
        evento = request.form["evento"]
        
        # Salva os dados na sessão, para uso posterior após a compra
        session['nome'] = nome
        session['email'] = email
        session['telefone'] = telefone
        session['evento'] = evento
        
        # Redireciona para a página de pagamento
        return redirect("/pagamento")
    
    return render_template("inscricao.html")


@app.route("/pagamento")
def pagamento():
    try:
        link_iniciar_pagamento = gerar_link_pagamento()
        return render_template("pagamento.html", link_pagamento=link_iniciar_pagamento)
    except Exception as e:
        print(f"Erro ao gerar link de pagamento: {e}")
        return render_template("erro.html", mensagem="Erro ao gerar link de pagamento.")
    

@app.route("/compracerta")
def compra_certa():
    try:
        # Confirma o pagamento bem-sucedido
        registrar_pagamento()  # Atualiza o contador no JSON

        # Pega os dados da sessão
        nome = session.get('nome')
        email = session.get('email')
        telefone = session.get('telefone')
        evento = session.get('evento')

        # Verifica se os dados existem na sessão antes de salvar no Excel
        if nome and email and telefone and evento:
            # Salva os dados no Excel
            salvar_no_excel(nome, email, telefone, evento)

        # Limpa os dados da sessão após salvar no Excel
        session.pop('nome', None)
        session.pop('email', None)
        session.pop('telefone', None)
        session.pop('evento', None)

        return render_template("compracerta.html")
    except Exception as e:
        print(f"Erro ao registrar pagamento: {e}")
        return render_template("erro.html", mensagem="Erro ao registrar pagamento.")

@app.route("/compraerrada")
def compra_errada():
    return render_template("compraerrada.html")

if __name__ == "__main__":
    inicializar_json()  # Inicializa o arquivo JSON, se necessário
    app.run(debug=True)
